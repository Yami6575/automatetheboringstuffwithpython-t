import sys
import openpyxl
import os

from openpyxl.utils import get_column_letter




n_table = 4

wb = openpyxl.Workbook()
sheet = wb.active


for r in range(2, n_table + 2):
    sheet['A' + str(r)] = r - 1
    col_let = get_column_letter(r)
    sheet[col_let + str(1)] = r - 1
    


for col in range(2, n_table + 2):
    for row in range(2, n_table + 2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col - 1) * (row - 1)

wb.save('Multiplication Table.xlsx')